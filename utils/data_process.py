import numpy as np
import pandas as pd
import time
import matplotlib.pyplot as plt
from pylab import *
from openpyxl import load_workbook
import xlrd
#数据统计处理
origin_data_filename = "../data/fault.txt"
csv_file_name = '../data/cars/1004.csv'
class DataProcess(object):
    TIMEFORMAT = "%Y-%m-%d %H:%M:%S"
    DATAITEM = ["time", "car_code", "fault_code", "is_master", "fault_pattern", "fault_state"]
    DATAITEMINDEX = {"is_master": 0, "fault_code": 1, "car_code": 2, "time": 3, "fault_pattern": 4, "fault_state": 5}

    def __init__(self, filename, data_item=DATAITEM, generate_csv=True, csv_filename=csv_file_name, zero_v_set=set(), notice_class_fault_set=set()):
        if generate_csv:
            self.origin_data = self.readtxt(filename, data_item, zero_v_set=zero_v_set, notice_class_fault_set=notice_class_fault_set)
            self.to_csv(csv_filename)
        else:
            self.origin_data = pd.read_csv(csv_filename)
            print(self.origin_data)

    def convert_time(self, timestamp):
        loca_time = time.localtime(int(timestamp))
        return time.strftime(DataProcess.TIMEFORMAT, loca_time)

    def convert_timestamp(self, timestr):
        timeArray = time.strptime(timestr, "%Y-%m-%d %H:%M:%S")
        return int(time.mktime(timeArray))

    def get_carriage(self, car_code):
        return str(car_code)[-2:]

    def get_car(self, car_code):
        return str(car_code)[:-2]

    def readtxt(self, filename, data_item, zero_v_set, notice_class_fault_set):
        cou = 0
        with open(filename) as f:
            line = f.readline()
            data_item_nums = len(data_item)
            data_list = [[] for _ in range(data_item_nums)]
            while line:
                line = line[4:]
                per_line_list = line.split('|')
                data_nums = len(per_line_list)
                for i in range(7, data_nums):
                    fault = per_line_list[i]
                    data_info_list = fault.split(',')
                    vertifycode = data_info_list[2][:-2] + data_info_list[3][:-1]
                    if vertifycode in zero_v_set or data_info_list[1] in notice_class_fault_set:
                        continue
                    for j in range(data_item_nums):
                        index = DataProcess.DATAITEMINDEX[data_item[j]]
                        data_list[j].append(data_info_list[index].strip())
                cou = cou + 1
                line = f.readline()
        print(cou)
        csv_content = {}
        for i in range(data_item_nums):
            csv_content[data_item[i]] = data_list[i]
        dataframe = pd.DataFrame(csv_content)
        dataframe['time'] = dataframe['time'].apply(lambda x: str(int(x) // 10 * 10))
        dataframe = dataframe.sort_values(by='time')
        dataframe['date'] = dataframe['time'].apply(lambda x: self.convert_time(x))
        dataframe['car'] = dataframe['car_code'].apply(lambda x : x[:-2])
        dataframe.drop_duplicates(subset=["time", "fault_code", "car"], inplace=True)
        print(dataframe)
        return dataframe

    def to_csv(self, target_file_name='../data/test.csv'):
        self.origin_data.to_csv(target_file_name, index=False, sep=',')

    def divide_by_car_code(self):
        self.origin_data.insert(self.origin_data.shape[1], 'carriage', self.origin_data['car_code'].apply(lambda x: self.get_carriage(x)))
        self.origin_data['car_code'] = self.origin_data['car_code'].apply(lambda x: self.get_car(x))
        self.car_group_data = self.origin_data.groupby('car_code')
        cou = 0
        Max_nums = 0
        t = ""
        nums_list = []
        car_code_list = []
        for car_code, group in self.car_group_data:
            out_put_filename = '../data/cars/' + car_code + '.csv'
            print(car_code)
            print(group)
            nums_list.append(len(group))
            car_code_list.append(car_code)
            if(len(group) > Max_nums):
                Max_nums, t = len(group), car_code
            group.to_csv(out_put_filename, index=False, sep=',')
        df = pd.DataFrame({'car_code': car_code_list, 'nums': nums_list})
        df = df.sort_values(by='nums', ascending=False)
        df = df[:10]
        print(df)
        print(Max_nums, t)

    def count_nums_by_time(self, time_interval=30):
        time_interval = time_interval * 60
        self.origin_data['time'] = self.origin_data['date'].apply(lambda x: self.convert_timestamp(x))
        time_serials = self.origin_data['time'].values.tolist()
        serials_len = len(time_serials)
        self.counts_serial = []
        self.time_serial = []
        i = 0
        start_time = time_serials[0]
        end_time = time_serials[0] + time_interval
        while i < serials_len:
            if time_serials[-1] < end_time:
                break
            cou = 0
            j = i
            while(time_serials[j] - start_time <= time_interval):
                cou = cou + 1
                j = j + 1
            self.counts_serial.append(cou)
            self.time_serial.append(start_time)
            start_time = end_time
            end_time = end_time + time_interval
            i = j
        print(len(self.counts_serial))
        print(self.counts_serial)
        target_csv_filename = 'counts_serial.csv'
        df = pd.DataFrame({'time': self.time_serial, 'fault_nums': self.counts_serial})
        df['time'] = df['time'].apply(lambda x: self.convert_time(x))
        df.to_csv(target_csv_filename, index=False, sep=',')
        plt.plot(df['time'], df['fault_nums'])
        plt.show()

    def count_nums_by_day(self):
        self.origin_data['day'] = self.origin_data['time'].apply(lambda x: "%s" % x[:10])
        self.nums_by_days = self.origin_data.groupby('day').size()
        self.nums_by_days.to_csv("../data/1004_nums_by_day.csv")
        # df['Week/Year'] = df['Timestamp'].apply(lambda x: "%d/%d" % (x.week, x.year))

class BigDataProcess(object):
    fault_path = "E:/TieKeYan/data/fault"
    parameter_path = "E:/TieKeYan/data/parameter"
    gps_path = "E:/TieKeYan/data/gps"
    self_check_path = "E:/TieKeYan/data/selfcheck"
    def __init__(self, txtfilename):
        self.txtfilename = txtfilename
    def split_fault(self):
        cou = 0
        filename = 'fault.txt'
        with open(self.txtfilename, 'r') as f:
            line = f.readline()
            while line:
                if line[:2] == '01' or line[:2] == '02':
                    with open(filename, 'a+') as f1:
                        f1.write(line)
                cou = cou + 1
                line = f.readline()
            print(cou)
    def split_diff_data(self):
        cou = [0] * 6
        filename = "temp.txt"
        num_dict = {1, 2, 3, 5}
        with open(self.txtfilename, 'r') as f:
            line = f.readline()
            while line:
                g = int(line[1])
                print(g)
                if g in num_dict and cou[g] < 3:
                    with open(filename, 'a+') as f1:
                        f1.write(line)
                        print(g)
                        cou[g] = cou[g] + 1
                line = f.readline()
                if cou[1] > 2 and cou[2] > 2 and cou[3] > 2 and cou[5] > 2:
                    break



    def split_by_car_code(self):
        with open(self.txtfilename, 'r') as f:
            line = f.readline()
            cou = 1
            target_code = '2065'
            while line:
                per_line_list = line[4:].split('|')
                car_code = per_line_list[5]
                flag = True
                if line[:2] == '03' and car_code == target_code:
                    filename = self.parameter_path + "/" + target_code
                elif (line[:2] == '01' or line[:2] == '02') and car_code == target_code:
                    filename = self.fault_path + "/" + target_code
                elif line[:2] == '04' and car_code == target_code:
                    filename = self.self_check_path + "/" + target_code
                elif line[:2] == '05' and car_code == target_code:
                    filename = self.gps_path + "/" + target_code
                else:
                    flag = False
                if flag:
                    filename = filename + '.txt'
                    with open(filename, 'a+') as f1:
                        f1.write(line)
                cou = cou + 1
                line = f.readline()
            print(cou)

    def extracte_parameter(self, needed_parameters):
        target_file = '/2065.txt'
        parameter_filename = self.parameter_path + target_file
        cou = 0
        parameter_dict = dict()
        unuse_time = set()
        with open(parameter_filename, 'r') as f:
            line = f.readline()
            while line:
                cou = cou + 1
                line = line[4:]
                line = line[:-1]
                # print(line)
                line_list = line.split('|')
                item_time = int(line_list[6])
                if item_time in unuse_time:
                    line = f.readline()
                    continue
                if item_time in parameter_dict.keys():
                    temp_item = parameter_dict[item_time]
                else:
                    temp_item = dict()
                for i in range(7, len(line_list)):
                    parameter_info = line_list[i]
                    parameter_info_list = parameter_info.split(",")
                    if parameter_info_list[1] in needed_parameters:
                        temp_item[parameter_info_list[1]] = int(parameter_info_list[2])
                # print(temp_item)
                if "lcsd" in temp_item.keys() and temp_item["lcsd"] == 0:
                    unuse_time.add(item_time)
                    if item_time in parameter_dict:
                        del parameter_dict[item_time]
                else:
                    if temp_item:
                        parameter_dict[item_time] = temp_item
                line = f.readline()
        print(cou)
        print(parameter_dict)
        time_list = []
        xfwd_3_6_list = []
        xfwd_2_7_list = []
        xfwd_1_8_list = []
        xfwd_4_5_list = []
        lcsd_list = []
        for key in parameter_dict:
            time_list.append(key)
            temp = parameter_dict[key]
            if "3/6cxfwd" in temp.keys():
                xfwd_3_6_list.append(temp["3/6cxfwd"])
            else:
                xfwd_3_6_list.append(np.nan)
            if "2/7cxfwd" in temp.keys():
                xfwd_2_7_list.append(temp["2/7cxfwd"])
            else:
                xfwd_2_7_list.append(np.nan)
            if "1/8cxfwd" in temp.keys():
                xfwd_1_8_list.append(temp["1/8cxfwd"])
            else:
                xfwd_1_8_list.append(np.nan)
            if "4/5cxfwd" in temp.keys():
                xfwd_4_5_list.append(temp["4/5cxfwd"])
            else:
                xfwd_4_5_list.append(np.nan)
            if "lcsd" in temp.keys():
                lcsd_list.append(temp["lcsd"])
            else:
                lcsd_list.append(np.nan)

        df = pd.DataFrame({"time": time_list, "xfwd_1_8": xfwd_1_8_list, "xfwd_2_7": xfwd_2_7_list, "xfwd_3_6": xfwd_3_6_list, "xfwd_4_5": xfwd_4_5_list, "lcsd":lcsd_list})
        df = df.sort_values(by='time')
        df.to_csv("parameter.csv", index=False)

    def read_target_parameter(self):
        xfwd_record_file = self.parameter_path + "/xfwd.txt"
        speed_record_file = self.parameter_path + "/speed.txt"
        target_file = "E:/TieKeYan/113-66.txt"
        parameter_dict = dict()
        unuse_time = set()
        cou = 0
        needed_parameters = {"1/8cxfwd", "2/7cxfwd", "3/6cxfwd", "4/5cxfwd", "lcsd"}
        with open(target_file, 'r') as f:
            line = f.readline()
            while line:
                cou = cou + 1
                if line[0 : 2] != "03":
                    line = f.readline()
                    continue
                line = line[4:]
                line = line[:-1]
                # print(line)
                line_list = line.split('|')
                item_time_carcode = int(line_list[6] + line_list[5])
                if item_time_carcode in unuse_time:
                    line = f.readline()
                    continue
                if item_time_carcode in parameter_dict.keys():
                    temp_item = parameter_dict[item_time_carcode]
                else:
                    temp_item = dict()
                for i in range(7, len(line_list)):
                    parameter_info = line_list[i]
                    parameter_info_list = parameter_info.split(",")
                    if parameter_info_list[1] in needed_parameters:
                        temp_item[parameter_info_list[1]] = int(float(parameter_info_list[2]))
                # print(temp_item)
                if "lcsd" in temp_item.keys() and temp_item["lcsd"] == 0:
                    unuse_time.add(item_time_carcode)
                    if item_time_carcode in parameter_dict:
                        del parameter_dict[item_time_carcode]
                else:
                    if temp_item:
                        parameter_dict[item_time_carcode] = temp_item
                line = f.readline()
        print(cou)
        time_list = []
        xfwd_3_6_list = []
        xfwd_2_7_list = []
        xfwd_1_8_list = []
        xfwd_4_5_list = []
        lcsd_list = []
        for key in parameter_dict:
            time_list.append(key)
            temp = parameter_dict[key]
            if "3/6cxfwd" in temp.keys():
                xfwd_3_6_list.append(temp["3/6cxfwd"])
            else:
                xfwd_3_6_list.append(np.nan)
            if "2/7cxfwd" in temp.keys():
                xfwd_2_7_list.append(temp["2/7cxfwd"])
            else:
                xfwd_2_7_list.append(np.nan)
            if "1/8cxfwd" in temp.keys():
                xfwd_1_8_list.append(temp["1/8cxfwd"])
            else:
                xfwd_1_8_list.append(np.nan)
            if "4/5cxfwd" in temp.keys():
                xfwd_4_5_list.append(temp["4/5cxfwd"])
            else:
                xfwd_4_5_list.append(np.nan)
            if "lcsd" in temp.keys():
                lcsd_list.append(temp["lcsd"])
            else:
                lcsd_list.append(np.nan)

            df = pd.DataFrame(
                {"time": time_list, "xfwd_1_8": xfwd_1_8_list, "xfwd_2_7": xfwd_2_7_list, "xfwd_3_6": xfwd_3_6_list,
                 "xfwd_4_5": xfwd_4_5_list, "lcsd": lcsd_list})
            df = df.sort_values(by='time')
            df_part = df.iloc[0:100000]
            df_part.to_csv("part_parameter.csv", index=False)
            df.to_csv("Allparameter.csv", index=False)



def read_prompt_fault(file_name):
    workbook = load_workbook(filename=file_name)
    print(workbook.sheetnames)
    sheet = workbook["总表"]
    print(sheet)
    cell = sheet['AG172:AG1665']
    cou = 0
    s = set()
    for i in cell:
        for j in i:
            if j.value == '提示类':
                s.add(j.row)
    cell1 = sheet['D172:D1665']
    res = set()
    for i in cell1:
        for j in i:
            if j.row in s:
                res.add(j.value)
    return res

def read_fault_class(filename):
    workbook = load_workbook(filename=filename)
    print(workbook.sheetnames)
    sheet = workbook["CR400AF-A&B故障字典02.00大版本"]
    cell = sheet['A2:A1645']
    cell1 = sheet['B2:B1645']
    class_dict = {}
    class_code_dict = {}
    class_code_dict['正常'] = 0
    idx = 1
    for i in range(2, 1646):
        c1 = sheet.cell(row=i, column=1)
        c2 = sheet.cell(row=i, column=2)
        class_dict[str(c1.value)] = c2.value
        if  c2.value not in class_code_dict.keys():
            class_code_dict[c2.value] = idx
            idx = idx + 1
    print(class_dict)
    print(class_code_dict)
    return class_dict, class_code_dict


def get_zero_V(file_name):
    res = set()
    data = pd.read_csv(file_name)
    carno_time_list = data['carno&time']
    lcsd_list = data['lcsd']
    for i in range(len(lcsd_list)):
        if lcsd_list[i] == 0:
            temp = str(carno_time_list[i])
            res.add(temp[:-1])
    print(len(res))
    print(res)
    return res

def remove_0_v(filename):
    data = pd.read_csv(filename)
    print(data)
    zero_v_list = []
    lcsd_list = data['lcsd']
    for i in range(len(lcsd_list)):
        if lcsd_list[i] == 0:
            zero_v_list.append(i)
    data1 = data.drop(index=zero_v_list)
    print(data1)
    data1['carcode'] = data1['carno&time'].apply(lambda x: int(str(x)[0: 4]))
    data1['time'] = data1['carno&time'].apply(lambda x: int(str(x)[4: ]))
    data1 = data1.drop(['carno&time'], axis=1)
    order = ['time', 'carcode', 'lcsd']
    data1 = data1[order]
    data1 = data1.sort_values(by='time')
    data1.to_csv("speed.csv", index=False)

def remove_abnormal_t(filename):
    data = pd.read_csv(filename)
    print(data)
    abnormal_t_list = []
    xfwd_list = data['xfwd']
    for i in range(len(xfwd_list)):
        if xfwd_list[i] == -50:
            abnormal_t_list.append(i)
    data1 = data.drop(index=abnormal_t_list)
    print(data1)
    data1['carcode'] = data1['carno&time'].apply(lambda x: int(str(x)[0: 4]))
    data1['time'] = data1['carno&time'].apply(lambda x: int(str(x)[4:]))
    data1 = data1.drop(['carno&time'], axis=1)
    order = ['time', 'carcode', 'xfwd']
    data1 = data1[order]
    data1 = data1.sort_values(by='time')
    data1.to_csv("xfwd.csv", index=False)

def remove_abnormal_gps(filename):
    data = pd.read_csv(filename)
    print(data)
    abnormal_gps_list = []
    E_list = data['E']
    N_list = data['N']
    lcsd_list = data['lcsd']
    for i in range(len(E_list)):
        if E_list[i] == 0 or N_list[i] == 0:
            abnormal_gps_list.append(i)
    data1 = data.drop(index=abnormal_gps_list)
    print(data1)
    data1['carcode'] = data1['carno&time'].apply(lambda x: int(str(x)[0: 4]))
    data1['time'] = data1['carno&time'].apply(lambda x: int(str(x)[4:]))
    data1 = data1.drop(['carno&time'], axis=1)
    order = ['time', 'carcode', 'lcsd', 'E', 'N']
    data1 = data1[order]
    data1 = data1.sort_values(by='time')
    data1.to_csv("gps.csv", index=False)


def get_car_fault(carcode, notice_class_fault_set):
    cou = 0
    class_dict, class_code_dict = read_fault_class("CR400AF-A&B故障字典及故障报警字典V1.xlsx")
    zero_v_set = get_zero_V("../211223Res/spd.txt")
    print(class_dict)
    with open("fault.txt") as f:
        line = f.readline()
        data_item_nums = 5
        data_list = [[] for _ in range(data_item_nums)]
        while line:
            line = line[4:]
            per_line_list = line.split('|')
            if per_line_list[5] != carcode:
                line = f.readline()
                continue
            data_nums = len(per_line_list)
            for i in range(7, data_nums):
                fault = per_line_list[i]
                data_info_list = fault.split(',')
                vertifycode = data_info_list[2][:-2] + data_info_list[3][:-1]
                if  vertifycode in zero_v_set or data_info_list[1] in notice_class_fault_set:
                    continue
                if data_info_list[1] not in class_dict.keys():
                    continue
                fault_code = data_info_list[1].strip()
                fault_class = class_dict[fault_code]
                fault_class_code = class_code_dict[fault_class]
                data_list[0].append(data_info_list[3].strip())
                data_list[1].append(data_info_list[2].strip())
                data_list[2].append(data_info_list[1].strip())
                data_list[3].append(fault_class)
                data_list[4].append(fault_class_code)
            cou = cou + 1
            line = f.readline()
    print(cou)
    csv_content = {}
    csv_content["time"] = data_list[0]
    csv_content["car_code"] = data_list[1]
    csv_content["fault_code"] = data_list[2]
    csv_content["fault_class"] = data_list[3]
    csv_content["fault_class_code"] = data_list[4]
    dataframe = pd.DataFrame(csv_content)
    dataframe = dataframe.sort_values(by='time')
    print(dataframe)
    dataframe.to_csv("2065fault.csv", index=False)
    return dataframe

def get_all_fault(notice_class_fault_set):
    cou = 0
    class_dict, class_code_dict = read_fault_class("CR400AF-A&B故障字典及故障报警字典V1.xlsx")
    zero_v_set = get_zero_V("../211223Res/spd.txt")
    print(class_dict)
    with open("fault.txt") as f:
        line = f.readline()
        data_item_nums = 5
        data_list = [[] for _ in range(data_item_nums)]
        while line:
            line = line[4:]
            per_line_list = line.split('|')
            data_nums = len(per_line_list)
            for i in range(7, data_nums):
                fault = per_line_list[i]
                data_info_list = fault.split(',')
                vertifycode = data_info_list[2][:-2] + data_info_list[3][:-1]
                if  vertifycode in zero_v_set or data_info_list[1] in notice_class_fault_set:
                    continue
                if data_info_list[1] not in class_dict.keys():
                    continue
                fault_code = data_info_list[1].strip()
                fault_class = class_dict[fault_code]
                fault_class_code = class_code_dict[fault_class]
                data_list[0].append(data_info_list[3].strip())
                data_list[1].append(data_info_list[2].strip())
                data_list[2].append(data_info_list[1].strip())
                data_list[3].append(fault_class)
                data_list[4].append(fault_class_code)
            cou = cou + 1
            line = f.readline()
    print(cou)
    csv_content = {}
    csv_content["time"] = data_list[0]
    csv_content["car_code"] = data_list[1]
    csv_content["fault_code"] = data_list[2]
    csv_content["fault_class"] = data_list[3]
    csv_content["fault_class_code"] = data_list[4]
    dataframe = pd.DataFrame(csv_content)
    dataframe = dataframe.sort_values(by='time')
    print(dataframe)
    dataframe.to_csv("allfault.csv", index=False)
    return dataframe


def remove_same(filename, targetfilename):
    df = pd.read_csv(filename)
    print(df)
    df['time'] = df['time'].apply(lambda x: x // 10 * 10)
    df.drop_duplicates(subset=["time", "fault_code"], inplace=True)
    print(df)
    df.to_csv(targetfilename, index=False)

def convert_time(timestamp):
    loca_time = time.localtime(int(timestamp))
    return time.strftime(DataProcess.TIMEFORMAT, loca_time)

def make_multi_series(parameter_file, fault_file):
    df = pd.read_csv(parameter_file)
    df = df.iloc[47:12176]
    print(df)
    xfwd_index = df[["xfwd_1_8", "xfwd_2_7", "xfwd_3_6", "xfwd_4_5"]]
    time_dict = {}
    gps_df = pd.read_csv("2065gps.csv")
    Province_Count_dict = {'广东省':0, '安徽省':0, '湖北省':0, '湖南省':0, '江西省':0}
    Province_Fault_dict = {'广东省': 0, '安徽省': 0, '湖北省': 0, '湖南省': 0, '江西省': 0}
    df['E'] = 0.0
    df['N'] = 0.0
    for i in range(47, 12176):
        time_dict[df["time"][i]] = i
        time_point = df["time"][i]
        l = 0
        r = 14000
        while l < r:
            mid = (l + r) // 2
            if gps_df["time"][mid] < time_point:
                l = mid + 1
            else:
                r = mid
        Province_Count_dict[gps_df['Province'][r]] = Province_Count_dict[gps_df['Province'][r]] + 1
        df['E'][i] = gps_df['E'][r]
        df['N'][i] = gps_df['N'][r]
    print(Province_Count_dict)
    df["xfwd"] = xfwd_index.mean(axis=1)
    df["fault_class"] = 0
    df["fault_state"] = "正常"
    df["fault_code"] = "0"
    fault_df = pd.read_csv(fault_file)
    fault_df = fault_df.iloc[4:54]
    cou = 0
    lcsd_list = []
    xfwd_list = []
    Province_list = []
    gps_lcsd_list = []
    E_list = []
    N_list =[]
    for i in range(4, 54):
        if fault_df["time"][i] not in time_dict.keys():
            continue
        time_point = fault_df["time"][i]
        l = len(gps_df)
        k = 0
        while gps_df['time'][k] < time_point:
            k = k + 1
        Province_list.append(gps_df['Province'][k])
        Province_Fault_dict[gps_df['Province'][k]] = Province_Fault_dict[gps_df['Province'][k]] + 1
        E_list.append(gps_df['E'][k])
        N_list.append(gps_df['N'][k])
        gps_lcsd_list.append(gps_df['lcsd'][k])
        j = time_dict[fault_df["time"][i]]
        while df["fault_class"][j] != 0:
            j = j + 1
        df["fault_class"][j] = fault_df["fault_class_code"][i]
        df["fault_code"][j] = fault_df["fault_code"][i]
        df["fault_state"][j] = fault_df["fault_class"][i]
        lcsd_list.append(df["lcsd"][j])
        xfwd_list.append(df["xfwd"][j])
        cou = cou + 1
    print(cou)
    print(df)
    print(lcsd_list)
    print(xfwd_list)
    print(Province_list)
    print(gps_lcsd_list)
    print(Province_Fault_dict)
    print(E_list)
    print(N_list)
    df['date'] = df['time'].apply(lambda x: convert_time(x))
    df.to_csv("multi_series.csv", index=False)

def read_gps(filename, targetcode):
    df = pd.read_csv(filename)
    print(df)
    df["car"] = df['carno&time'].apply(lambda x: str(x)[:4])
    df["time"] = df['carno&time'].apply(lambda x: int(str(x)[4:]))
    df = df.loc[df['car'] == targetcode]
    df = df.drop(['carno&time'], axis=1)
    df = df.sort_values(by='time')
    df['date'] = df['time'].apply(lambda x: convert_time(x))
    print(df)
    df.to_csv("2065gps.csv", index=False)
    Province_set = set()
    for row in df.itertuples():
        Province_set.add(getattr(row, 'Province'))
    print(Province_set)
    return df

def merge_parameter():
    df_xfwd = pd.read_csv('xfwd.csv')
    abnormal_v_list = []
    t_list = df_xfwd['time']
    for i in range(len(t_list)):
        if t_list[i] < 1000000000:
            abnormal_v_list.append(i)
    df_xfwd = df_xfwd.drop(index=abnormal_v_list)
    df_xfwd.to_csv("xfwd.csv", index=False)

def remove_abnormal_time():
    df = pd.read_csv("speed.csv")
    df_clear = df.drop(df[df['time'] < 1600000000].index)
    print(df_clear)
    df_clear.to_csv("true_speed.csv", index=False)

def convert_file():
    df = pd.read_csv("multi_series.csv")
    df = df[['time', 'lcsd', 'E', 'N', 'xfwd']]
    df.to_csv('train_env.csv', index=False)

def get_integer_info(filename):
    workbook = xlrd.open_workbook(filename)
    worksheet = workbook.sheet_by_index(0)
    nrows = worksheet.nrows
    print(nrows)
    ncols = worksheet.ncols
    print(ncols)
    NAME = 1
    DATATYPE = 3
    DEVCODE = 11
    FAULTCLASS = 17
    uint_target_info = []
    int_target_info = []
    cc_target_info = []
    tmbin_target_info = []
    bcd_target_info = []
    type_nums = dict()
    data_class_info = dict()
    for i in range(1, nrows):
        cell_data_type = worksheet.cell_value(i, DATATYPE)
        if cell_data_type not in type_nums:
            type_nums[cell_data_type] = 1
        else:
            type_nums[cell_data_type] = type_nums[cell_data_type] + 1
        if cell_data_type == 'UINT':
            fault_class = worksheet.cell_value(i, FAULTCLASS)
            uint_target_info.append({'name': worksheet.cell_value(i, NAME), 'dev_code': worksheet.cell_value(i, DEVCODE), 'fault_class': fault_class})
            if fault_class not in data_class_info:
                data_class_info[fault_class] = 1
            else:
                data_class_info[fault_class] = data_class_info[fault_class] + 1
        if cell_data_type == 'INT':
            int_target_info.append({'name': worksheet.cell_value(i, NAME), 'dev_code': worksheet.cell_value(i, DEVCODE), 'fault_class': worksheet.cell_value(i, FAULTCLASS)})
        if cell_data_type == 'cc':
            cc_target_info.append({'name': worksheet.cell_value(i, NAME), 'dev_code': worksheet.cell_value(i, DEVCODE), 'fault_class': worksheet.cell_value(i, FAULTCLASS)})
        if cell_data_type == 'TMBIN':
            tmbin_target_info.append({'name': worksheet.cell_value(i, NAME), 'dev_code': worksheet.cell_value(i, DEVCODE), 'fault_class': worksheet.cell_value(i, FAULTCLASS)})
        if cell_data_type == 'BCD':
            bcd_target_info.append({'name': worksheet.cell_value(i, NAME), 'dev_code': worksheet.cell_value(i, DEVCODE), 'fault_class': worksheet.cell_value(i, FAULTCLASS)})
    print(uint_target_info)
    np.save("../data_record/uint_parameter_info.npy", uint_target_info)
    return uint_target_info

def genater_spd_csv(filename):
    res = set()
    data = pd.read_csv(filename)
    carno_time_list = data['carno&time']
    lcsd_list = data['lcsd']
    for i in range(len(lcsd_list)):
        if lcsd_list[i] == 0:
            temp = str(carno_time_list[i])
            res.add(temp[:-1])
    print(len(res))
    print(len(lcsd_list))
    np.save("../data_record/zero_v_info.npy", res)
    # print(res)
    return res


if __name__ == "__main__":
    bigdataprocess = BigDataProcess("E:/TieKeYan/113-66.txt")
    bigdataprocess.split_diff_data()
    # genater_spd_csv("../211223Res/spd.txt")
    # get_integer_info('CR400AF-A变量解析配置文件V1.0.xls')

    # get_zero_V("../211223Res/spd.txt")
    # bigdataprocess = BigDataProcess("E:/TieKeYan/113-66.txt")
    # bigdataprocess.split_fault()
    # convert_file()
    # bigdataprocess = BigDataProcess("E:/TieKeYan/113-66.txt")
    # bigdataprocess.read_target_parameter()
    # bigdataprocess.split_fault()
    # needed_parameters = {"1/8cxfwd", "2/7cxfwd", "3/6cxfwd", "4/5cxfwd", "lcsd"}
    # bigdataprocess.extracte_parameter(needed_parameters)

    #data_item_list = ['time', 'fault_code', 'car_code']
    #zero_v_set = get_zero_V("../211223Res/spd.txt")
    #notice_class_fault_set = read_prompt_fault("副本附件1.8：动车组诊断代码总表-CR400AF平台.xlsx")
    #get_car_fault("2065", notice_class_fault_set)
    # read_fault_class("CR400AF-A&B故障字典及故障报警字典V1.xlsx")
    #data_process = DataProcess("fault.txt", data_item_list, generate_csv=True, csv_filename="fault_info.csv", zero_v_set=zero_v_set, notice_class_fault_set=notice_class_fault_set)
    #data_process.count_nums_by_time(time_interval=15)
    # read_prompt_fault("副本附件1.8：动车组诊断代码总表-CR400AF平台.xlsx")
    # get_zero_V("../211223Res/spd.txt")
    # remove_same("2065fault.csv", "2065target.csv")
    # make_multi_series("parameter.csv", "2065target.csv")
    # read_gps("target.txt", "2065")

    # data_item_list = ['time', 'fault_code', 'car_code']
    # zero_v_set = get_zero_V("../211223Res/spd.txt")
    # notice_class_fault_set = read_prompt_fault("副本附件1.8：动车组诊断代码总表-CR400AF平台.xlsx")
    # get_all_fault(notice_class_fault_set)

    # remove_same("allfault.csv", "alltarget.csv")
    # remove_0_v("../211223Res/spd.txt")
    # remove_abnormal_t("../211223Res/temp.txt")
    # remove_abnormal_gps('../211223Res/gps.txt')
    # merge_parameter()
    # remove_abnormal_time()

    # df = pd.read_csv("../utils/multi_series.csv")
    # print(df)
    # df = df[["time", "lcsd", "E", "N", "xfwd"]]
    # print(df)
    # df.to_csv('train_fault.csv', sep=",", index=False)